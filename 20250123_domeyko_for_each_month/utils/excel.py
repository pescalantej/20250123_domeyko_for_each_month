# %%
import os
import pandas as pd
from typing import Dict
from io import BytesIO
from pathlib import WindowsPath, Path
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
# %%
def create_and_open_workbook(full_path: WindowsPath) -> Workbook:
    """
    Deletes the existing workbook if it exists, creates a new workbook, and returns the workbook object.

    Parameters:
    file_path (WindowsPath): The path to the directory containing the Excel file.
    file_name (str): The name of the Excel file.

    Returns:
    Workbook: The instance of the new Excel workbook.
    """
    
    if full_path.exists():
        # Delete the existing workbook
        os.remove(full_path)
        print(f"Deleted existing workbook: {full_path.name}")
    
    # Create a new workbook
    wb = Workbook()
    # Save the new workbook to the specified path
    wb.save(str(full_path))
    print(f"Created new workbook: {full_path.name}")
    return wb

# %%
def write_dataframe_to_sheet(
    wb: Workbook, 
    df: pd.DataFrame, 
    ws_name: str
    ) -> Worksheet:
    """
    Saves the given DataFrame to a specified sheet in the workbook.

    Parameters:
    wb (Workbook): The openpyxl workbook object.
    df (pd.DataFrame): The DataFrame to save.
    ws_name (str): The name of the sheet where the DataFrame will be saved.

    Returns:
    Worksheet: The newly created or updated worksheet.
    """
    # Check if the sheet already exists; if so, remove it
    if ws_name in wb.sheetnames:
        std = wb[ws_name]
        wb.remove(std)

    # Create a new sheet with the given name
    sheet = wb.create_sheet(title=ws_name)

    # Write the DataFrame to the new sheet
    for row in dataframe_to_rows(df, index=True, header=True):
        sheet.append(row)
    
    sheet.delete_rows(2)
    sheet.cell(row=1, column=1).value = "date"    
    
    print(f"DataFrame written to sheet {ws_name}.")
    
    return sheet

# %%
def delete_default_sheet(wb: Workbook):
    """
    Deletes the sheet with the name "Sheet" or "Hoja1" from an existing workbook if it exists.
    
    Parameters:
    wb (Workbook): The openpyxl Workbook object from which the sheet will be deleted.
    
    Returns:
    None
    
    Prints a message indicating whether the sheet was deleted or if it did not exist.
    """
    # Check if the sheet exists in the workbook
    default_sheets = ["Sheet", "Hoja1"]
    for sheet_name in default_sheets:
        if sheet_name in wb.sheetnames:
            # Delete the specified sheet
            sheet_to_delete = wb[sheet_name]
            wb.remove(sheet_to_delete)
            print(f"Sheet '{sheet_name}' deleted.")

# %%
def set_font_size(ws: Worksheet, font_size: int):
    """
    Sets the font size of all cells in the given sheet to the specified font size.

    Parameters:
    ws (Worksheet): The openpyxl Worksheet object.
    font_size (int): The desired font size to set.
    """
    # Iterate through all cells in the sheet and set font size
    for row in ws.iter_rows():
        for cell in row:
            # Set the font size for each cell
            cell.font = Font(size=font_size)

# %%
def set_row_height(ws: Worksheet, row_number: int, height: float):
    """
    Sets the height of the specified row in the given sheet to the specified height.

    Parameters:
    ws (Worksheet): The openpyxl Worksheet object.
    row_number (int): The row number to set the height for.
    height (float): The desired height to set for the specified row.
    """
    # Set the row height for the specified row number
    ws.row_dimensions[row_number].height = height

# %%
def set_full_grid(ws: Worksheet):
    """
    Sets gridlines (borders) around the data range in the given sheet.
    
    Parameters:
    ws (Worksheet): The openpyxl Worksheet object in which gridlines will be set around the data range.
    
    Returns:
    None
    """
    # Define a border style (thin line)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Determine the used range in the sheet
    min_row = ws.min_row
    min_col = ws.min_column
    max_row = ws.max_row
    max_col = ws.max_column

    # Iterate through each cell in the used range and apply the border
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = thin_border

# %%
def auto_adjust_and_align_first_column(ws: Worksheet) -> None:
    """
    Automatically adjusts the width of the first column of the given sheet and sets the alignment to left.

    Parameters:
    ws (Worksheet): An openpyxl Worksheet object.

    Returns:
    None
    """
    max_length = 0

    # Iterate through each cell in the first column ('A') to find the maximum length
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            if cell.value:
                # Calculate the length of the cell value and add a small padding
                cell_length = len(str(cell.value))
                max_length = max(max_length, cell_length)

                # Set the cell alignment to left
                cell.alignment = Alignment(horizontal='left')

    # Set the width of the first column ('A') based on the maximum content length
    adjusted_width = max_length + 1
    ws.column_dimensions['A'].width = adjusted_width

# %%
def set_column_width_except_first(ws: Worksheet, width: float, adjustment_factor: float = 0.78):
    """
    Sets the width of the columns around the data range, except the first column, in the given sheet.
    
    Parameters:
    ws (Worksheet): The openpyxl Worksheet object in which the column widths will be set.
    width (float): The desired width for the columns.
    adjustment_factor (float, optional): The adjustment factor to apply to the width. Defaults to 0.78.
    
    Returns:
    None
    """

    # Get the maximum column number in the sheet
    max_col = ws.max_column

    # Set the width for columns except the first one
    for col in range(2, max_col + 1):  # Start from column 2 (B) to max column
        col_letter = ws.cell(row=1, column=col).column_letter
        ws.column_dimensions[col_letter].width = width + adjustment_factor

# %%
def set_row_bold(ws: Worksheet, row_number: int):
    """
    Sets the specified row of the given sheet to bold.

    Parameters:
    ws (Worksheet): The openpyxl Worksheet object in which the specified row will be set to bold.
    row_number (int): The row number to set as bold (1-based index).

    Returns:
    None
    """
    # Iterate through each cell in the specified row and set the font to bold
    for cell in ws[row_number]:
        cell.font = Font(bold=True)

# %%
def top_left_alignment_and_wrap_text_first_row(ws: Worksheet) -> None:
    """
    Sets the alignment of the first row of the given sheet to top-left and wraps the text.

    Parameters:
    ws (Worksheet): An openpyxl Worksheet object.

    Returns:
    None
    """
    # Iterate through each cell in the first row of the sheet
    for cell in ws[1]:
        # Set horizontal alignment to left, vertical alignment to top, and wrap text
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

# %%
def set_format_01(ws: Worksheet) -> None:
    """
    Applies a set of formatting operations to the workbook. The operations are as follows:

    1. Sets the font size of all cells in all sheets to 10.
    2. Sets the height of the first row in all sheets to 30.
    3. Sets gridlines (borders) around the data range for all sheets.
    4. Sets the first column of each sheet to auto-adjust its width.
    5. Sets the width of the columns around the data range, except the first column, to 12.
    6. Sets the font of the first row in each sheet to bold.
    7. Aligns and wraps the text of the first row in each sheet.
    Parameters:
    wb (Workbook): The openpyxl Workbook object to format.

    Returns:
    None
    """
    set_font_size(ws, 10) # it takes to long
    set_row_height(ws, 1, 45)
    set_full_grid(ws) # it takes to long
    set_column_width_except_first(ws, 12)
    set_row_bold(ws, 1)
    auto_adjust_and_align_first_column(ws)
    top_left_alignment_and_wrap_text_first_row(ws)

    print(f"Worksheet {ws.title} formatted with format_01.")

# %%
def insert_dataframe_into_template(
        template_path: Path,
        data: Dict[str, pd.DataFrame]
        ) -> BytesIO:
    """
    Updates an Excel file template with data from a dictionary of dataframes and returns the file as a BytesIO stream.

    Args:
        template_path (Path): Path to the Excel template file to be used.
        data (Dict[str, pd.DataFrame]): A dictionary where keys are sheet names and values are dataframes to insert.

    Returns:
        BytesIO: A stream containing the updated Excel file.

    Raises:
        FileNotFoundError: If the template file is not found.
        KeyError: If a sheet name in the data dictionary is not found in the template file.
        ValueError: If data contains invalid dataframes.
    """
    # Load the Excel template
    try:
        workbook = load_workbook(template_path)
    except FileNotFoundError:
        raise FileNotFoundError(f"The template file '{template_path}' does not exist.")

    # Process each sheet and update with corresponding dataframe
    for sheet_name, dataframe in data.items():
        if sheet_name not in workbook.sheetnames:
            raise KeyError(f"Sheet '{sheet_name}' not found in the template.")
        if not isinstance(dataframe, pd.DataFrame):
            raise ValueError(f"Value for sheet '{sheet_name}' must be a pandas DataFrame.")

        sheet = workbook[sheet_name]
        dataframe_reset = dataframe.reset_index(names="Date")
        dataframe_no_style = dataframe_reset.copy()

        # Write the new dataframe to the sheet
        for r_idx, row in enumerate(dataframe_to_rows(dataframe_no_style, index=False, header=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                sheet.cell(row=r_idx, column=c_idx, value=value)

    # Save the workbook to a BytesIO stream
    output_stream = BytesIO()
    workbook.save(output_stream)
    output_stream.seek(0)

    return output_stream